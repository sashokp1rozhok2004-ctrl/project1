from pathlib import Path

import pandas as pd
from django.utils import timezone
from openpyxl import load_workbook


def validate_uploaded_file(uploaded_file):
    """
    Выполняет базовую техническую проверку Excel или CSV.
    Изменяет поля проверки у объекта UploadedFile и сохраняет их.
    """
    uploaded_file.validation_status = "not_checked"
    uploaded_file.sheet_count = 0
    uploaded_file.row_count = 0
    uploaded_file.column_count = 0
    uploaded_file.validation_message = ""

    try:
        file_path = uploaded_file.file.path
        extension = Path(file_path).suffix.lower()

        if extension == ".xlsx":
            result = validate_xlsx(file_path)

        elif extension == ".xls":
            result = validate_xls(file_path)

        elif extension == ".csv":
            result = validate_csv(file_path)

        else:
            raise ValueError("Формат файла не поддерживается.")

        uploaded_file.validation_status = result["status"]
        uploaded_file.sheet_count = result["sheet_count"]
        uploaded_file.row_count = result["row_count"]
        uploaded_file.column_count = result["column_count"]
        uploaded_file.validation_message = result["message"]

    except Exception as error:
        uploaded_file.validation_status = "error"
        uploaded_file.validation_message = (
            f"Не удалось прочитать файл: {error}"
        )

    uploaded_file.validated_at = timezone.now()

    uploaded_file.save(
        update_fields=[
            "validation_status",
            "sheet_count",
            "row_count",
            "column_count",
            "validation_message",
            "validated_at",
        ]
    )


def validate_xlsx(file_path):
    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True,
    )

    sheet_count = len(workbook.sheetnames)
    total_rows = 0
    max_columns = 0
    warnings = []

    for sheet in workbook.worksheets:
        non_empty_rows = 0
        non_empty_columns = 0
        has_data = False

        for row in sheet.iter_rows(values_only=True):
            values = list(row)

            if any(value not in (None, "") for value in values):
                has_data = True
                non_empty_rows += 1

                last_filled_column = 0

                for index, value in enumerate(values, start=1):
                    if value not in (None, ""):
                        last_filled_column = index

                non_empty_columns = max(
                    non_empty_columns,
                    last_filled_column,
                )

        if not has_data:
            warnings.append(
                f'Лист «{sheet.title}» полностью пустой.'
            )

        total_rows += non_empty_rows
        max_columns = max(max_columns, non_empty_columns)

    workbook.close()

    if total_rows == 0:
        return {
            "status": "error",
            "sheet_count": sheet_count,
            "row_count": 0,
            "column_count": 0,
            "message": "В файле не найдено данных.",
        }

    if warnings:
        return {
            "status": "warning",
            "sheet_count": sheet_count,
            "row_count": total_rows,
            "column_count": max_columns,
            "message": "\n".join(warnings),
        }

    return {
        "status": "valid",
        "sheet_count": sheet_count,
        "row_count": total_rows,
        "column_count": max_columns,
        "message": "Файл успешно прочитан.",
    }


def validate_xls(file_path):
    workbook = pd.ExcelFile(file_path)

    sheet_count = len(workbook.sheet_names)
    total_rows = 0
    max_columns = 0
    warnings = []

    for sheet_name in workbook.sheet_names:
        dataframe = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
        )

        dataframe = dataframe.dropna(how="all")

        if dataframe.empty:
            warnings.append(
                f'Лист «{sheet_name}» полностью пустой.'
            )
            continue

        total_rows += len(dataframe.index)
        max_columns = max(
            max_columns,
            len(dataframe.columns),
        )

    if total_rows == 0:
        return {
            "status": "error",
            "sheet_count": sheet_count,
            "row_count": 0,
            "column_count": 0,
            "message": "В файле не найдено данных.",
        }

    status = "warning" if warnings else "valid"
    message = (
        "\n".join(warnings)
        if warnings
        else "Файл успешно прочитан."
    )

    return {
        "status": status,
        "sheet_count": sheet_count,
        "row_count": total_rows,
        "column_count": max_columns,
        "message": message,
    }


def validate_csv(file_path):
    read_attempts = [
        {"encoding": "utf-8", "sep": None},
        {"encoding": "utf-8-sig", "sep": None},
        {"encoding": "cp1251", "sep": None},
    ]

    last_error = None
    dataframe = None

    for options in read_attempts:
        try:
            dataframe = pd.read_csv(
                file_path,
                engine="python",
                **options,
            )
            break

        except Exception as error:
            last_error = error

    if dataframe is None:
        raise ValueError(
            f"CSV не удалось прочитать: {last_error}"
        )

    dataframe = dataframe.dropna(how="all")

    if dataframe.empty:
        return {
            "status": "error",
            "sheet_count": 1,
            "row_count": 0,
            "column_count": 0,
            "message": "CSV-файл не содержит данных.",
        }

    return {
        "status": "valid",
        "sheet_count": 1,
        "row_count": len(dataframe.index),
        "column_count": len(dataframe.columns),
        "message": "Файл успешно прочитан.",
    }