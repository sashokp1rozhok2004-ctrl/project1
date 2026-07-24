from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


MONTH_NAMES = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def normalize_text(value: Any) -> str:
    """
    Приводит значение Excel к удобному текстовому виду.
    """

    if value is None:
        return ""

    text = str(value)
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = re.sub(r"\s+", " ", text)

    return text.strip()

def normalize_article(value: Any) -> str:
    """
    Преобразует артикул в текст без научной записи.

    Например:
    9700000000.0 -> "9700000000"
    """

    if value is None:
        return ""

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))

        return format(value, "f").rstrip("0").rstrip(".")

    return normalize_text(value)

def parse_month(value: Any) -> date | None:
    """
    Пытается определить месяц и год из ячейки Excel.

    Поддерживает:
    - настоящую дату Excel;
    - 'Январь 2024';
    - 'январь 2024'.
    """

    if isinstance(value, datetime):
        return date(value.year, value.month, 1)

    if isinstance(value, date):
        return date(value.year, value.month, 1)

    text = normalize_text(value).lower()

    if not text:
        return None

    year_match = re.search(r"\b(20\d{2})\b", text)

    if not year_match:
        return None

    year = int(year_match.group(1))

    for month_name, month_number in MONTH_NAMES.items():
        if month_name in text:
            return date(year, month_number, 1)

    return None


def is_total_row(
    article: Any,
    product_name: Any,) -> bool:
    """
    Определяет строки итогов и служебные строки,
    которые не должны попадать в результат.
    """

    article_text = normalize_text(article).lower()
    product_text = normalize_text(product_name).lower()

    total_words = {
            "итого",
            "всего",
            "общий итог",
    }

    return (
            article_text in total_words
            or product_text in total_words
            or article_text.startswith("итого ")
            or product_text.startswith("итого ")
    )

def normalize_metric(value: Any) -> str | None:
    """
    Определяет тип показателя в пятой строке.
    """

    text = normalize_text(value).lower()

    if not text:
        return None

    if "aop" in text or "аор" in text:
        return "aop"

    if "прогноз" in text:
        return "forecast"

    if "факт" in text:
        return "actual"

    if "коммент" in text:
        return "comment"

    return None


def find_main_sheet(workbook) -> str:
    """
    Ищет лист, похожий на основную таблицу менеджера.
    """

    required_headers = {
        "клиент",
        "менеджер",
        "наименование",
    }

    for worksheet in workbook.worksheets:
        row_four = {
            normalize_text(worksheet.cell(4, column).value).lower()
            for column in range(1, worksheet.max_column + 1)
        }

        has_required_headers = all(
            any(required in cell for cell in row_four)
            for required in required_headers
        )

        has_article = any(
            "артикул" in cell or "номер изделия" in cell
            for cell in row_four
        )

        row_five = [
            normalize_metric(worksheet.cell(5, column).value)
            for column in range(1, worksheet.max_column + 1)
        ]

        has_month_metrics = (
            "aop" in row_five
            or "forecast" in row_five
            or "actual" in row_five
        )

        if has_required_headers and has_article and has_month_metrics:
            return worksheet.title

    raise ValueError(
        "Не удалось автоматически определить основной лист."
    )


def build_month_columns(worksheet) -> list[dict[str, Any]]:
    """
    Определяет месячные столбцы по строкам 4 и 5.

    В строке 4 месяц обычно указан только у первого столбца блока.
    Поэтому найденный месяц переносится вправо на следующие показатели.
    """

    month_columns: list[dict[str, Any]] = []
    current_month: date | None = None

    for column in range(1, worksheet.max_column + 1):
        month_value = worksheet.cell(4, column).value
        metric_value = worksheet.cell(5, column).value

        parsed_month = parse_month(month_value)

        if parsed_month is not None:
            current_month = parsed_month

        metric = normalize_metric(metric_value)

        if current_month is not None and metric is not None:
            month_columns.append(
                {
                    "column": column,
                    "month": current_month,
                    "metric": metric,
                }
            )

    if not month_columns:
        raise ValueError(
            "В таблице не найдены месячные столбцы."
        )

    return month_columns


def find_base_columns(worksheet) -> dict[str, int]:
    """
    Ищет номера основных служебных столбцов.
    """

    aliases = {
        "client": ["клиент"],
        "payment_terms": [
            "условия платежа",
            "условия оплаты",
        ],
        "supplier": ["поставщик"],
        "manager": ["менеджер"],
        "row_number": ["п/п", "п\\п"],
        "product_flag": ["продукт wabco"],
        "article": [
            "номер изделия",
            "артикул",
        ],
        "product_name": ["наименование"],
    }

    found: dict[str, int] = {}

    for column in range(1, worksheet.max_column + 1):
        header = normalize_text(
            worksheet.cell(4, column).value
        ).lower()

        for field_name, possible_names in aliases.items():
            if field_name in found:
                continue

            if any(name in header for name in possible_names):
                found[field_name] = column

    required = [
        "client",
        "manager",
        "article",
        "product_name",
    ]

    missing = [
        field_name
        for field_name in required
        if field_name not in found
    ]

    if missing:
        raise ValueError(
            "Не найдены обязательные столбцы: "
            + ", ".join(missing)
        )

    return found


def read_manager_file(
    file_path: str | Path,
    manager_name: str | None = None,
    forecast_date: date | None = None,
) -> pd.DataFrame:
    """
    Разбирает один менеджерский Excel-файл.

    Возвращает таблицу, где один ряд соответствует:
    товару + клиенту + месяцу.
    """

    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(
            f"Файл не найден: {file_path}"
        )

    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True,
    )

    sheet_name = find_main_sheet(workbook)
    worksheet = workbook[sheet_name]

    base_columns = find_base_columns(worksheet)
    month_columns = build_month_columns(worksheet)

    records: list[dict[str, Any]] = []

    last_values: dict[str, Any] = {
        "client": None,
        "payment_terms": None,
        "supplier": None,
        "manager": None,
    }

    for row in range(6, worksheet.max_row + 1):
        base_values: dict[str, Any] = {}

        for field_name, column in base_columns.items():
            value = worksheet.cell(row, column).value
            base_values[field_name] = value

        for field_name in last_values:
            value = base_values.get(field_name)

            if value not in (None, ""):
                last_values[field_name] = value
            else:
                base_values[field_name] = last_values[field_name]

        article = normalize_article(
            base_values.get("article")
        )

        product_name = normalize_text(
            base_values.get("product_name")
        )

        if article in (None, "") and product_name in (None, ""):
            continue

        if is_total_row(article, product_name):
            continue

        resolved_manager = (
            manager_name
            or normalize_text(base_values.get("manager"))
            or "Не определён"
        )

        month_data: dict[date, dict[str, Any]] = {}

        for month_info in month_columns:
            month = month_info["month"]
            metric = month_info["metric"]
            column = month_info["column"]

            month_data.setdefault(
                month,
                {
                    "aop": None,
                    "forecast": None,
                    "actual": None,
                    "comment": None,
                },
            )

            month_data[month][metric] = (
                worksheet.cell(row, column).value
            )

        for month, metrics in month_data.items():
            has_any_value = any(
                value not in (None, "")
                for value in metrics.values()
            )

            if not has_any_value:
                continue

            records.append(
                {
                    "manager": resolved_manager,
                    "client": base_values.get("client"),
                    "payment_terms": base_values.get(
                        "payment_terms"
                    ),
                    "supplier": base_values.get("supplier"),
                    "row_number": base_values.get("row_number"),
                    "product_flag": base_values.get(
                        "product_flag"
                    ),
                    "article": article,
                    "product_name": product_name,
                    "month": month,
                    "aop": metrics["aop"],
                    "forecast": metrics["forecast"],
                    "actual": metrics["actual"],
                    "comment": metrics["comment"],
                    "forecast_date": forecast_date,
                    "source_file": file_path.name,
                    "source_sheet": sheet_name,
                    "source_row": row,
                }
            )

    workbook.close()

    dataframe = pd.DataFrame(records)

    if dataframe.empty:
        raise ValueError(
            "После обработки не найдено ни одной строки данных."
        )

    return dataframe

